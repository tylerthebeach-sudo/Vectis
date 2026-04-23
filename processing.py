"""
processing.py — JobTime export → Sage 300 Timecard import
"""
import os
import pandas as pd
import logging
import shutil
import tempfile
from datetime import datetime, date
from config import AppConfig
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Full column specs — exact match to Sage 300 import template ───────────────

HEADER_COLS = [
    'EMPLOYEE', 'PEREND', 'TIMECARD', 'TCARDDESC',
    'TIMESLATE', 'REUSECARD', 'ACTIVE', 'SEPARATECK', 'PROCESSED',
    'CREGHRS', 'CSHIFTHRS', 'CVACHRSP', 'CVACHRSA', 'CSICKHRSP', 'CSICKHRSA',
    'CCOMPHRSP', 'CCOMPHRSA', 'CVACAMTP', 'CVACAMTA', 'CSICKAMTP', 'CSICKAMTA',
    'CCOMPAMTP', 'CCOMPAMTA', 'CDISIHRSP', 'CDISIHRSA', 'CDISIAMTP', 'CDISIAMTA',
    'LASTNAME', 'FIRSTNAME', 'MIDDLENAME',
    'GREGHRS', 'GSHIFTHRS', 'GVACHRSP', 'GVACHRSA', 'GSICKHRSP', 'GSICKHRSA',
    'GCOMPHRSP', 'GCOMPHRSA', 'GVACAMTP', 'GVACAMTA', 'GSICKAMTP', 'GSICKAMTA',
    'GCOMPAMTP', 'GCOMPAMTA', 'KEYACTION',
    'GDISIHRSP', 'GDISIHRSA', 'GDISIAMTP', 'GDISIAMTA',
    'HIREDATE', 'FIREDATE', 'PARTTIME', 'PAYFREQ', 'OTSCHED', 'COMPTIME',
    'SHIFTSCHED', 'SHIFTNUM', 'WORKPROV', 'STATUS', 'INACTDATE', 'PROCESSCMD',
    'GOTHOURS', 'OTCALCTYPE', 'HRSPERDAY', 'WORKCODE', 'TOTALJOBS', 'USERSEC',
    'WKLYFLSA', 'VALUES', 'OTOVERRIDE', 'COTHOURS', 'TCDLINES', 'SWJOB', 'SRCEAPPL',
]

DETAIL_COLS = [
    'EMPLOYEE', 'PEREND', 'TIMECARD', 'LINENUM',
    'CATEGORY', 'EARNDED', 'EARDEDTYPE', 'EARDEDDATE', 'STARTTIME', 'STOPTIME',
    'GLSEG1', 'GLSEG2', 'GLSEG3',
    'HOURS', 'CALCMETH', 'LIMITBASE', 'CNTBASE', 'RATE',
    'PAYORACCR', 'EXPACCT', 'LIABACCT', 'OTACCT', 'SHIFTACCT', 'ASSETACCT',
    'OTSCHED', 'SHIFTSCHED', 'SHIFTNUM', 'WCC', 'TAXWEEKS', 'TAXANNLIZ',
    'WEEKLYNTRY', 'ENTRYTYPE', 'POOLEDTIPS', 'DESC',
    'GLSEGID1', 'GLSEGDESC1', 'GLSEGID2', 'GLSEGDESC2', 'GLSEGID3', 'GLSEGDESC3',
    'KEYACTION', 'WORKPROV', 'PROCESSCMD', 'NKEMPLOYEE', 'NKPEREND', 'NKTIMECARD',
    'NKLINENUM', 'DAYS', 'WCCGROUP', 'VALUES', 'OTHOURS', 'OTRATE', 'SWFLSA',
    'DISTCODE', 'REXPACCT', 'RLIABACCT', 'SWALLOCJOB', 'JOBS', 'WORKCODE',
    'JOBHOURS', 'JOBBASE', 'RCALCMETH', 'RLIMITBASE', 'RRATEOVER', 'RRATE', 'DEFRRATE',
]

OPTIONAL_FIELD_VALUES_COLS = [
    'EMPLOYEE', 'PEREND', 'TIMECARD', 'OPTFIELD', 'VALUE', 'TYPE', 'LENGTH',
    'DECIMALS', 'ALLOWNULL', 'VALIDATE', 'SWSET', 'KEYACTION', 'VALINDEX',
    'VALIFTEXT', 'VALIFMONEY', 'VALIFNUM', 'VALIFLONG', 'VALIFBOOL',
    'VALIFDATE', 'VALIFTIME', 'FDESC', 'VDESC',
]

DETAILS_OPTIONAL_FIELD_COLS = [
    'EMPLOYEE', 'PEREND', 'TIMECARD', 'LINENUM', 'OPTFIELD', 'VALUE', 'TYPE',
    'LENGTH', 'DECIMALS', 'ALLOWNULL', 'VALIDATE', 'SWSET', 'KEYACTION',
    'VALINDEX', 'VALIFTEXT', 'VALIFMONEY', 'VALIFNUM', 'VALIFLONG', 'VALIFBOOL',
    'VALIFDATE', 'VALIFTIME', 'FDESC', 'VDESC', 'CATEGORY', 'EARNDED', 'ENTRYTYPE',
]

JOB_DETAILS_COLS = [
    'EMPLOYEE', 'PEREND', 'TIMECARD', 'LINENUM', 'JOBLINE', 'CONTRACT',
    'PROJECT', 'CCATEGORY', 'IDCUST', 'CURRCODE', 'STARTTIME', 'STOPTIME',
    'HOURS', 'CNTBASE', 'BILLTYPE', 'BILLRATE', 'ARITEMNO', 'ARUNIT',
    'WIPACCT', 'OTWIPACCT', 'STWIPACCT', 'OTHOURS', 'OTBILLRATE', 'VALUES',
    'PROJSTYLE', 'PROJTYPE', 'UFMTCONTNO', 'REVREC', 'COSTCLASS', 'CURRDESC',
    'KEYACTION', 'ALLOWGLOVR', 'PROCESSCMD', 'RESOURCE', 'RESDESC',
]

JOBS_OPTIONAL_FIELD_COLS = [
    'EMPLOYEE', 'PEREND', 'TIMECARD', 'LINENUM', 'JOBLINE', 'OPTFIELD',
    'VALUE', 'TYPE', 'LENGTH', 'DECIMALS', 'ALLOWNULL', 'VALIDATE', 'SWSET',
    'KEYACTION', 'VALINDEX', 'VALIFTEXT', 'VALIFMONEY', 'VALIFNUM', 'VALIFLONG',
    'VALIFBOOL', 'VALIFDATE', 'VALIFTIME', 'FDESC', 'VDESC',
]

# ── Fixed values per Sage 300 template ────────────────────────────────────────

HEADER_FIXED = {
    'TIMESLATE':  0,      'REUSECARD':  False,  'ACTIVE':     True,
    'SEPARATECK': False,  'PROCESSED':  False,
    'CREGHRS':    0,      'CSHIFTHRS':  0,
    'CVACHRSP':   0,      'CVACHRSA':   0,      'CSICKHRSP':  0,    'CSICKHRSA':  0,
    'CCOMPHRSP':  0,      'CCOMPHRSA':  0,      'CVACAMTP':   0,    'CVACAMTA':   0,
    'CSICKAMTP':  0,      'CSICKAMTA':  0,      'CCOMPAMTP':  0,    'CCOMPAMTA':  0,
    'CDISIHRSP':  0,      'CDISIHRSA':  0,      'CDISIAMTP':  0,    'CDISIAMTA':  0,
    'MIDDLENAME': '',
    'GREGHRS':    0,      'GSHIFTHRS':  0,
    'GVACHRSP':   0,      'GVACHRSA':   0,      'GSICKHRSP':  0,    'GSICKHRSA':  0,
    'GCOMPHRSP':  0,      'GCOMPHRSA':  0,      'GVACAMTP':   0,    'GVACAMTA':   0,
    'GSICKAMTP':  0,      'GSICKAMTA':  0,      'GCOMPAMTP':  0,    'GCOMPAMTA':  0,
    'KEYACTION':  0,
    'GDISIHRSP':  0,      'GDISIHRSA':  0,      'GDISIAMTP':  0,    'GDISIAMTA':  0,
    'PARTTIME':   False,  'PAYFREQ':    4,       'SHIFTNUM':   0,
    'WORKPROV':   1,      'STATUS':     1,       'PROCESSCMD': 0,
    'GOTHOURS':   0,      'OTCALCTYPE': 0,       'HRSPERDAY':  8,
    'TOTALJOBS':  0,      'USERSEC':    0,       'WKLYFLSA':   False,
    'VALUES':     0,      'OTOVERRIDE': False,   'COTHOURS':   0,
    'SWJOB':      False,  'SRCEAPPL':   'CP',
}

DETAIL_FIXED = {
    'CATEGORY':   2,      'EARDEDTYPE': 1,       'STARTTIME':  0,    'STOPTIME':   0,
    'CALCMETH':   4,      'LIMITBASE':  0,       'CNTBASE':    0,    'PAYORACCR':  6,
    'SHIFTNUM':   0,      'WCC':        'A',     'TAXWEEKS':   0,    'TAXANNLIZ':  0,
    'WEEKLYNTRY': False,  'ENTRYTYPE':  1,       'POOLEDTIPS': 0,    'KEYACTION':  0,
    'WORKPROV':   1,      'PROCESSCMD': 0,       'NKLINENUM':  0,    'DAYS':       0,
    'WCCGROUP':   'W',    'VALUES':     0,       'OTHOURS':    0,    'OTRATE':     0,
    'SWFLSA':     False,  'SWALLOCJOB': False,   'JOBS':       0,    'JOBHOURS':   0,
    'JOBBASE':    0,      'RCALCMETH':  0,       'RLIMITBASE': 0,    'RRATEOVER':  False,
    'RRATE':      0,      'DEFRRATE':   0,
}


# ── Distribution mapping ───────────────────────────────────────────────────────

def load_distribution_map(path: str) -> dict:
    """
    Load employee → GL distribution mapping from the distribution Excel file.

    Returns dict keyed by employee ID string:
      { emp_id: {'distcode': str, 'gl_acct': int|None, 'subf_acct': int|None} }

    gl_acct  → EXPACCT/OTACCT/SHIFTACCT for all earn codes except SUBF
    subf_acct → EXPACCT/OTACCT/SHIFTACCT when EARNDED == 'SUBF'
    """
    if not path:
        return {}
    if not os.path.exists(path):
        raise ValueError(f"Distribution file not found:\n{path}")
    try:
        ext = path.lower().rsplit('.', 1)[-1]
        engine = 'xlrd' if ext == 'xls' else 'openpyxl'
        df = pd.read_excel(path, engine=engine)

        dist_map = {}
        for _, row in df.iterrows():
            try:
                raw_id = str(row['EMPLOYEE'])
                emp_id = raw_id.split('.')[0].strip()
                if not emp_id or emp_id == 'nan':
                    continue

                gl_raw   = row.get('GL Distribution')
                subf_raw = row.get('SUBF')

                gl_acct   = int(float(str(gl_raw)))   if pd.notna(gl_raw)   else None
                subf_acct = int(float(str(subf_raw))) if pd.notna(subf_raw) else None

                dist_map[emp_id] = {
                    'distcode':  str(row['CLASS1']).strip(),
                    'gl_acct':   gl_acct,
                    'subf_acct': subf_acct,
                }
            except Exception:
                continue

        logging.info(f"Loaded distribution map: {len(dist_map)} employees from {path}")
        return dist_map
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Could not read distribution file:\n{e}")


# ── File reading ───────────────────────────────────────────────────────────────

def _read_df(path: str, ext: str) -> pd.DataFrame:
    if ext == 'csv':
        try:
            return pd.read_csv(path, encoding='utf-8', low_memory=False)
        except UnicodeDecodeError:
            return pd.read_csv(path, encoding='cp1252', low_memory=False)
    elif ext in ('xlsx', 'xlsm'):
        return pd.read_excel(path, sheet_name=0, engine='openpyxl')
    elif ext == 'xls':
        return pd.read_excel(path, sheet_name=0, engine='xlrd')
    else:
        raise ValueError(f"Unsupported file type: .{ext}")


def read_input_file(path: str) -> pd.DataFrame:
    ext = path.lower().rsplit('.', 1)[-1]
    try:
        df = _read_df(path, ext)
    except ValueError:
        raise
    except PermissionError:
        try:
            with tempfile.NamedTemporaryFile(suffix=f'.{ext}', delete=False) as tmp:
                tmp_path = tmp.name
            shutil.copy2(path, tmp_path)
            df = _read_df(tmp_path, ext)
        except Exception as e:
            raise ValueError(f"Could not read file (it may be open in another program):\n{e}")
    except Exception as e:
        raise ValueError(f"Could not read file:\n{e}")
    if df.empty:
        raise ValueError("The selected file is empty.")
    logging.info(f"Read {len(df)} rows from {path}")
    return df


def detect_columns(df: pd.DataFrame) -> dict:
    """Find the Employee, Pay Code, Hours, and Work Date columns."""
    aliases = {
        'employee':  ['Employee', 'Employee No', 'Employee Number', 'Emp No'],
        'paycode':   ['Pay Code', 'PayCode', 'Pay Type'],
        'hours':     ['Hours', 'Total Hours', 'Hrs'],
        'work_date': ['Work Date', 'WorkDate', 'Date'],
    }
    col_map = {}
    lc = {c.lower(): c for c in df.columns if isinstance(c, str)}
    for field, opts in aliases.items():
        col_map[field] = None
        for opt in opts:
            if opt in df.columns:
                col_map[field] = opt
                break
            if opt.lower() in lc:
                col_map[field] = lc[opt.lower()]
                break
    return col_map


def validate_input(df: pd.DataFrame, col_map: dict) -> list[str]:
    errors = []
    for field in ('employee', 'paycode', 'hours'):
        if not col_map.get(field):
            errors.append(f"Required column not found: '{field}'")
    if errors:
        return errors
    non_num = pd.to_numeric(df[col_map['hours']], errors='coerce').isna().sum()
    if non_num:
        errors.append(f"Warning: {non_num} rows have non-numeric Hours — they will be skipped.")
    return errors


def flag_anomalies(df: pd.DataFrame, col_map: dict) -> dict:
    """Returns {row_index: reason_string} for every row with a data problem."""
    problems = {}
    hours_col = col_map.get('hours')
    emp_col   = col_map.get('employee')
    pc_col    = col_map.get('paycode')

    if hours_col and hours_col in df.columns:
        mask = pd.to_numeric(df[hours_col], errors='coerce').isna()
        for idx in df.index[mask]:
            val = df.at[idx, hours_col]
            problems[idx] = (
                f"Non-numeric Hours value: \"{val}\"\n\n"
                f"This row will be skipped during export because Sage 300 "
                f"requires a number in the Hours field."
            )

    if emp_col and emp_col in df.columns:
        mask = df[emp_col].isna() | (df[emp_col].astype(str).str.strip() == '')
        for idx in df.index[mask]:
            if idx not in problems:
                problems[idx] = (
                    "Missing Employee value.\n\n"
                    "This row has no employee ID and will be skipped during export."
                )

    if pc_col and pc_col in df.columns:
        mask = df[pc_col].isna() | (df[pc_col].astype(str).str.strip() == '')
        for idx in df.index[mask]:
            if idx not in problems:
                problems[idx] = (
                    "Missing Pay Code value.\n\n"
                    "This row has no pay code and will be skipped during export."
                )

    return problems


def flag_dist_anomalies(df: pd.DataFrame, col_map: dict, dist_map: dict) -> dict:
    """
    Returns {row_index: reason_string} for every row whose employee ID is
    not present in the distribution map.  These rows block export entirely.
    """
    problems = {}
    emp_col = col_map.get('employee')
    if not emp_col or emp_col not in df.columns or not dist_map:
        return problems

    raw = df[emp_col].astype(str).str.strip()
    has_dash = raw.str.contains(' - ', regex=False)
    emp_ids = raw.where(~has_dash, raw.str.split(' - ').str[0].str.strip())

    valid   = (emp_ids != '') & (emp_ids != 'nan')
    missing = valid & ~emp_ids.isin(dist_map)

    for idx in df.index[missing]:
        emp_id = emp_ids.at[idx]
        problems[idx] = (
            f"Employee not in distribution file: \"{emp_id}\"\n\n"
            f"This employee has no distribution code (DISTCODE) or GL account mapping. "
            f"Without this, Sage 300 will not know how to categorize the payroll expense "
            f"and will reject the import file.\n\n"
            f"To fix:\n"
            f"  1. Add employee {emp_id} to the distribution file\n"
            f"     with their CLASS1 and GL Distribution values.\n"
            f"  2. Click \u2018Select Distribution File\u2026\u2019 to reload it."
        )
    return problems


def find_unmapped_paycodes(
    df: pd.DataFrame,
    col_map: dict,
    config,
    date_from=None,
    date_to=None,
) -> dict:
    """
    Returns which pay codes in the file have no enabled mapping and how many hours
    they represent, using the same cleaning/date-filter logic as process_timesheet.

    Result: {'unmapped': {paycode: hours}, 'total_hours': float, 'mapped_hours': float}
    """
    paycode_col = col_map.get('paycode')
    hours_col   = col_map.get('hours')
    emp_col     = col_map.get('employee')
    date_col    = col_map.get('work_date')

    if not paycode_col or not hours_col or not emp_col:
        return {'unmapped': {}, 'total_hours': 0.0, 'mapped_hours': 0.0}

    work = df.copy()
    work[hours_col] = pd.to_numeric(work[hours_col], errors='coerce')
    work = work.dropna(subset=[emp_col, paycode_col, hours_col])
    work = work[work[hours_col] > 0]

    if date_col and date_col in work.columns and (date_from or date_to):
        work[date_col] = pd.to_datetime(work[date_col], errors='coerce')
        if date_from:
            work = work[work[date_col].dt.date >= date_from]
        if date_to:
            work = work[work[date_col].dt.date <= date_to]

    if work.empty:
        return {'unmapped': {}, 'total_hours': 0.0, 'mapped_hours': 0.0}

    total_hours = round(float(work[hours_col].sum()), 2)
    enabled = {m['jobtime_code'].strip() for m in config.pay_code_mappings if m.get('enabled', True)}
    work['_paycode'] = work[paycode_col].astype(str).str.strip()

    mapped_hours = round(float(work[work['_paycode'].isin(enabled)][hours_col].sum()), 2)
    unmapped = (
        work[~work['_paycode'].isin(enabled)]
        .groupby('_paycode')[hours_col]
        .sum()
        .round(2)
        .to_dict()
    )
    return {'unmapped': unmapped, 'total_hours': total_hours, 'mapped_hours': mapped_hours}


def flag_unmapped_paycodes_rows(df: pd.DataFrame, col_map: dict, config) -> dict:
    """
    Returns {row_index: reason_string} for every row whose pay code has no enabled mapping.
    These rows are silently excluded during export.
    """
    paycode_col = col_map.get('paycode')
    if not paycode_col or paycode_col not in df.columns:
        return {}
    enabled = {m['jobtime_code'].strip() for m in config.pay_code_mappings if m.get('enabled', True)}
    pc = df[paycode_col].astype(str).str.strip()
    mask = pc.ne('nan') & pc.ne('') & ~pc.isin(enabled)
    return {
        idx: f'Pay code "{pc.at[idx]}" has no mapping — this row will be excluded from the export.'
        for idx in df.index[mask]
    }


def get_excluded_rows(
    df: pd.DataFrame,
    col_map: dict,
    config,
    date_from=None,
    date_to=None,
) -> pd.DataFrame:
    """
    Returns the subset of raw df rows that process_timesheet would silently drop
    because their pay code is not in the enabled mapping (after date filter).
    """
    paycode_col = col_map.get('paycode')
    hours_col   = col_map.get('hours')
    emp_col     = col_map.get('employee')
    date_col    = col_map.get('work_date')

    if not paycode_col or not hours_col or not emp_col:
        return pd.DataFrame()

    work = df.copy()
    work[hours_col] = pd.to_numeric(work[hours_col], errors='coerce')
    work = work.dropna(subset=[emp_col, paycode_col, hours_col])
    work = work[work[hours_col] > 0]

    if date_col and date_col in work.columns and (date_from or date_to):
        work[date_col] = pd.to_datetime(work[date_col], errors='coerce')
        if date_from:
            work = work[work[date_col].dt.date >= date_from]
        if date_to:
            work = work[work[date_col].dt.date <= date_to]

    if work.empty:
        return pd.DataFrame()

    enabled = {m['jobtime_code'].strip() for m in config.pay_code_mappings if m.get('enabled', True)}
    pc = work[paycode_col].astype(str).str.strip()
    excluded = work[~pc.isin(enabled)].copy()
    return excluded.drop(columns=[c for c in excluded.columns if c.startswith('_')], errors='ignore')


def _parse_employee(raw: str) -> tuple[str, str, str]:
    """Parse '20220064 - Thomson, Scott' into (emp_id, lastname, firstname)."""
    raw = str(raw).strip()
    if ' - ' in raw:
        parts  = raw.split(' - ', 1)
        emp_id = parts[0].strip()
        name   = parts[1].strip()
    else:
        return raw, '', ''

    if ',' in name:
        name_parts = name.split(',', 1)
        lastname   = name_parts[0].strip().upper()
        firstname  = name_parts[1].strip().upper()
    else:
        name_parts = name.split(' ', 1)
        firstname  = name_parts[0].strip().upper()
        lastname   = name_parts[1].strip().upper() if len(name_parts) > 1 else ''

    return emp_id, lastname, firstname


# ── Core processing ────────────────────────────────────────────────────────────

def process_timesheet(
    df:        pd.DataFrame,
    col_map:   dict,
    config:    AppConfig,
    date_from: date | None = None,
    date_to:   date | None = None,
    dist_map:  dict | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Transform a JobTime DataFrame into Sage 300 Header + Detail DataFrames.
    Returns (header_df, detail_df, summary_dict).
    """
    emp_col     = col_map['employee']
    paycode_col = col_map['paycode']
    hours_col   = col_map['hours']
    date_col    = col_map.get('work_date')

    # ── 1. Clean ──────────────────────────────────────────────────────────────
    work = df.copy()
    work[hours_col] = pd.to_numeric(work[hours_col], errors='coerce')
    work = work.dropna(subset=[emp_col, paycode_col, hours_col])
    work = work[work[hours_col] > 0]

    # ── 2. Date filter ────────────────────────────────────────────────────────
    if date_col and date_col in work.columns and (date_from or date_to):
        work[date_col] = pd.to_datetime(work[date_col], errors='coerce')
        if date_from:
            work = work[work[date_col].dt.date >= date_from]
        if date_to:
            work = work[work[date_col].dt.date <= date_to]

    if work.empty:
        raise ValueError("No valid rows after filtering. Check dates and file contents.")

    # ── 3. Filter to enabled pay codes ────────────────────────────────────────
    enabled = {m['jobtime_code'].strip() for m in config.pay_code_mappings if m.get('enabled', True)}
    work['_paycode'] = work[paycode_col].astype(str).str.strip()
    work = work[work['_paycode'].isin(enabled)]

    if work.empty:
        raise ValueError(
            "No rows match the enabled pay code mappings.\n"
            "Check that your Pay Code Mappings match the values in the file."
        )

    # ── 4. Resolve Sage EARNDED code ──────────────────────────────────────────
    pc_to_earnded = {
        m['jobtime_code'].strip(): m['earnded']
        for m in config.pay_code_mappings if m.get('enabled', True)
    }
    work['_earnded'] = work['_paycode'].map(pc_to_earnded)

    # ── 5. Parse employee ID + name ───────────────────────────────────────────
    parsed = work[emp_col].astype(str).apply(_parse_employee)
    work['_emp_id']    = parsed.apply(lambda x: x[0])
    work['_lastname']  = parsed.apply(lambda x: x[1])
    work['_firstname'] = parsed.apply(lambda x: x[2])

    # ── 6. Aggregate hours: one row per employee per earn code ────────────────
    agg = (
        work.groupby(['_emp_id', '_earnded', '_lastname', '_firstname'], sort=False)[hours_col]
        .sum()
        .reset_index()
        .rename(columns={hours_col: '_hours'})
    )
    agg['_hours'] = agg['_hours'].round(2)

    # ── 7. PEREND date ────────────────────────────────────────────────────────
    perend_str = config.get('perend_date', '').strip()
    try:
        perend = datetime.strptime(perend_str, '%Y-%m-%d')
    except ValueError:
        if date_col and date_col in work.columns:
            perend = pd.to_datetime(work[date_col], errors='coerce').max().to_pydatetime()
        else:
            perend = datetime.today()
        logging.warning(f"PEREND not set — defaulting to {perend.date()}")

    timecard_code = config.get('timecard_code', '4')
    timecard_desc = config.get('timecard_desc', '')

    # ── 8. Earn code sort order ───────────────────────────────────────────────
    earnded_order = [
        m['earnded'] for m in config.pay_code_mappings if m.get('enabled', True)
    ]
    # User-defined LINENUM per earn code; fall back to position * 1000 if not set
    earnded_linenum = {
        m['earnded']: int(m['linenum'])
        for i, m in enumerate(config.pay_code_mappings)
        if m.get('enabled', True) and m.get('linenum')
    }

    # ── 9. Build Detail rows ──────────────────────────────────────────────────
    detail_rows   = []
    missing_dist  = set()
    tcdlines_map  = {}   # emp_id → count of detail lines

    for emp_id, emp_group in agg.groupby('_emp_id', sort=False):
        emp_dist  = (dist_map or {}).get(emp_id, {})
        distcode  = emp_dist.get('distcode', '')
        gl_acct   = emp_dist.get('gl_acct')
        subf_acct = emp_dist.get('subf_acct')

        if not emp_dist:
            missing_dist.add(emp_id)

        # Sort earn lines by configured mapping order
        emp_group = emp_group.copy()
        emp_group['_order'] = emp_group['_earnded'].apply(
            lambda e: earnded_order.index(e) if e in earnded_order else 999
        )
        emp_group = emp_group.sort_values('_order')

        for fallback_idx, (_, earn_row) in enumerate(emp_group.iterrows(), start=1):
            earnded = earn_row['_earnded']
            linenum = earnded_linenum.get(earnded, fallback_idx * 1000)
            # SUBF earn code posts to a different GL account than all other codes
            acct = subf_acct if (earnded == 'SUBF' and subf_acct is not None) else gl_acct

            detail_rows.append({
                'EMPLOYEE':  emp_id,
                'PEREND':    perend,
                'TIMECARD':  timecard_code,
                'LINENUM':   linenum,
                'EARNDED':   earnded,
                'HOURS':     earn_row['_hours'],
                'EXPACCT':   acct,
                'OTACCT':    acct,
                'SHIFTACCT': acct,
                'DISTCODE':  distcode,
                **DETAIL_FIXED,
            })

        tcdlines_map[emp_id] = len(emp_group)

    detail_df = pd.DataFrame(detail_rows, columns=DETAIL_COLS)

    # ── 10. Build Header rows (one per employee) ──────────────────────────────
    employees = agg[['_emp_id', '_lastname', '_firstname']].drop_duplicates('_emp_id')
    header_rows = []
    for _, emp in employees.iterrows():
        header_rows.append({
            'EMPLOYEE':  emp['_emp_id'],
            'PEREND':    perend,
            'TIMECARD':  timecard_code,
            'TCARDDESC': timecard_desc,
            'LASTNAME':  emp['_lastname'],
            'FIRSTNAME': emp['_firstname'],
            'TCDLINES':  tcdlines_map.get(emp['_emp_id'], 0),
            **HEADER_FIXED,
        })

    header_df = pd.DataFrame(header_rows, columns=HEADER_COLS)

    summary = {
        'employee_count': len(employees),
        'total_rows_in':  len(df),
        'rows_processed': len(work),
        'total_hours':    round(float(agg['_hours'].sum()), 2),
        'perend':         perend.date(),
        'detail_lines':   len(detail_rows),
        'missing_dist':   sorted(missing_dist),
    }
    logging.info(f"Processed: {summary}")
    return header_df, detail_df, summary


# ── Excel export ───────────────────────────────────────────────────────────────

def export_to_excel(header_df: pd.DataFrame, detail_df: pd.DataFrame, output_path: str):
    """Write all 6 Sage 300 import sheets to an Excel workbook."""
    wb = Workbook()

    ws_h = wb.active
    ws_h.title = 'Timecard_Header'
    _write_sheet(ws_h, header_df)

    ws_d = wb.create_sheet('Timecard_Detail')
    _write_sheet(ws_d, detail_df)

    # Required empty sheets
    for sheet_name, cols in [
        ('Timecard_Optional_Field_Values',  OPTIONAL_FIELD_VALUES_COLS),
        ('Timecard_Details_Optional_Field', DETAILS_OPTIONAL_FIELD_COLS),
        ('Timecard_Job_Details',            JOB_DETAILS_COLS),
        ('Timecard_Jobs_Optional_Field_Va', JOBS_OPTIONAL_FIELD_COLS),
    ]:
        _write_sheet(wb.create_sheet(sheet_name), pd.DataFrame(columns=cols))

    wb.save(output_path)
    logging.info(f"Exported to {output_path}")


def _write_sheet(ws, df: pd.DataFrame):
    FILL  = PatternFill('solid', fgColor='1F3864')
    FONT  = Font(bold=True, color='FFFFFF', size=10)
    ALIGN = Alignment(horizontal='center', vertical='center')

    cols = list(df.columns)

    for c_idx, col in enumerate(cols, 1):
        cell            = ws.cell(row=1, column=c_idx, value=col)
        cell.font       = FONT
        cell.fill       = FILL
        cell.alignment  = ALIGN

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, col in enumerate(cols, 1):
            val  = getattr(row, col, None)
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(val, datetime):
                cell.value        = val
                cell.number_format = 'YYYY-MM-DD'
            elif val != val:  # NaN
                cell.value = ''
            else:
                cell.value = val

    for c_idx, col in enumerate(cols, 1):
        max_len = max(
            len(col),
            df[col].astype(str).str.len().max() if not df.empty else 0
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 30)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'
